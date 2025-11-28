"""

Translates mappings to .xlsx spreadsheets for validation

"""

from ast import pattern
import pandas as pd
import os
import time
import re
import openpyxl
import csv
import json

from tqdm import tqdm
from openpyxl.worksheet.datavalidation import DataValidation    
from collections import defaultdict

def mappings_to_spreadsheet_from_xml(mappings_directory: str, excel_file: str, all_mappings=True) -> None:
    os.system("CLS")
    print(f"Putting all mappings in {mappings_directory} into spreadsheet\n<waiting for 5 seconds before proceeding>") if all_mappings else None
    time.sleep(5) if all_mappings else None

    headers = ["Source", "Target", "Method", "Score", "Good Mapping?", "Suggested mapping", "Comments"]


    collate_by_method = {m: [] for m in ['kge', 'lightweight']}
    flag = verify_xlsx(excel_file)

    print(f"Excel file {excel_file} verified or created successfully.") if flag else print(f"Failed to verify or create Excel file {excel_file}.")

    for method in collate_by_method:
        ensure_sheet_for_method(excel_file, method)


    for mappings in tqdm(os.listdir(mappings_directory)):
        if mappings.endswith(".xml"):
            method = next((m for m in collate_by_method if m in mappings), None)

            if method:
                xmlfile = os.path.join(mappings_directory, mappings)
                matches = find_xml_mapping_score(xmlfile)

                # Extract the method name as the token between the 2nd and 3rd underscore
                base = os.path.splitext(mappings)[0]  # remove .xml
                parts = base.split("_")
                if len(parts) >= 3:
                    # token between 2nd and 3rd underscore (0-based index 2)
                    method_name = parts[2]
                else:
                    # fallback to previous behaviour if filename doesn't match expected pattern
                    method_name = base.replace("mappings", "").strip("_")
                for source, target, score in matches:
                    collate_by_method[method].append([source, target, method_name, score, "", ""])

    ## -- Create workbook or open existing
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    for method, rows in tqdm(
        {k: v for k, v in collate_by_method.items() if v}.items(),
        desc="Processing"
        ):
        
        if method not in wb.sheetnames:
            ws = wb.create_sheet(title=method)
            ws.append(["Source", "Target", "Method", "Score", "Good Mapping?", "Suggested mapping", "Comments"])
        else:
            ws = wb[method]
        for row in rows:
            ws.append(row)
        makedropdown(ws)
    

    wb.save(excel_file)
    wb.close()

def mappings_to_spreadsheet_from_tsv(tsv_dirs, excel_file, all_mappings=True):

    os.system("CLS")
    print(f"Putting all mappings in {tsv_dirs} into spreadsheet\n<waiting for 5 seconds before proceeding>") if all_mappings else None
    time.sleep(5) if all_mappings else None

    flag = verify_xlsx(excel_file)
    print(f"Excel file {excel_file} verified or created successfully.") if flag else print(f"Failed to verify or create Excel file {excel_file}.")

    ensure_sheet_for_method(excel_file, "logmap")

    wb = openpyxl.load_workbook(excel_file)
    ws = wb["logmap"]

    for tsv_file in tqdm(os.listdir(tsv_dirs), desc="Processing TSV files"):
        if tsv_file.endswith(".tsv"):
            with open(os.path.join(tsv_dirs, tsv_file), encoding="utf-8") as f:
                for line in f:
                    line = line.strip().strip('"')  # Remove newline and outer quotes
                    if not line or line.lower().startswith("src"):  # Skip header or empty
                        continue

                    parts = line.split('\t')
                    if len(parts) < 5:
                        continue  # Skip malformed lines

                    # Compose source and target with labels and <=> as requested
                    source = f"{parts[0]} ({parts[3]})"
                    target = f"{parts[1]} ({parts[4]})"
                    method = "logmap"
                    score = parts[2]
                    ws.append([source, target, method, score, "", ""])
            makedropdown(ws)

    wb.save(excel_file)
    wb.close()

def mappings_to_spreadsheet_from_json(json_file, excel_file, all_mappings=True):
    os.system("CLS")
    print(f"Putting all mappings in {json_file} into spreadsheet\n<waiting for 5 seconds before proceeding>") if all_mappings else None
    time.sleep(5) if all_mappings else None

    flag = verify_xlsx(excel_file)
    print(f"Excel file {excel_file} verified or created successfully.") if flag else print(f"Failed to verify or create Excel file {excel_file}.")

    ensure_sheet_for_method(excel_file, "logmap")

    wb = openpyxl.load_workbook(excel_file)
    ws = wb["bertmap"]

    with open(json_file, encoding="utf-8") as f:
        data = json.load(f)
        for item in data["training"]:
            pass

            

def makedropdown(ws):
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)

    ## -- Apply to all rows in column E (Good Mapping?)
    for row in range(2, ws.max_row + 1):
        dv.add(ws[f"E{row}"])

def find_xml_mapping_score(xmlfile):
    pattern = re.compile(
    r'<entity1 rdf:resource="([^"]+)"\/>\s*'
    r'<entity2 rdf:resource="([^"]+)"\/>\s*'
    r'<relation>=<\/relation>\s*'
    r'<measure [^>]*>([\d\.eE+-]+)</measure>',
    re.MULTILINE | re.DOTALL
)
    xml_text = open(xmlfile, encoding="utf-8").read()
    matches = pattern.findall(xml_text)

    return matches       

def ensure_sheet_for_method(excel_file, method_name):
    """
    Ensure a sheet for the given method exists in the Excel file.
    If it does not exist, create it.
    """
    wb = openpyxl.load_workbook(excel_file)
    if method_name not in wb.sheetnames:
        ws = wb.create_sheet(title=method_name)
        ## -- Add column headers
        ws.append(["Source", "Target", "Method", "Score", "Good Mapping?", "Comments"])
        makedropdown(ws)
        wb.save(excel_file)
    wb.close()

def verify_xlsx(excel_file) -> bool:
    # Ensure the parent directory exists
    parent_dir = os.path.dirname(excel_file)
    if not os.path.exists(parent_dir):
        os.makedirs(parent_dir)

    if os.path.exists(excel_file):
        try:
            wb = openpyxl.load_workbook(excel_file)
            wb.close()
            return True
        except Exception:
            os.remove(excel_file)  # Remove corrupted file
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Temp"
            wb.save(excel_file)
            wb.close()
            return True
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Temp"
        wb.save(excel_file)
        wb.close()
        return True
    


if __name__ == "__main__":
    tsvdirectory = os.path.join(os.getcwd(), "alignment", "outputs", "logmap")
    xmldirectory = os.path.join(os.getcwd(), "alignment", "outputs", "ontoaligner")
    excel_file = os.path.join(os.getcwd(), "alignment", "outputs", "summary.xlsx")

    mappings_to_spreadsheet_from_tsv(tsvdirectory, excel_file)
    mappings_to_spreadsheet_from_xml(xmldirectory, excel_file)